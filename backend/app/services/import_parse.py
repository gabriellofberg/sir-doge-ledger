from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..config import MAX_IMPORT_ROWS

_CURRENCY_RE = re.compile(
    r"(?i)(?:\s*(?:sek|kr|nok|dkk|eur|usd|gbp)\s*|€|\$|£)"
)


@dataclass
class ParsedRow:
    tx_date: str  # ISO date
    amount: float
    description: str


@dataclass
class ColumnMapping:
    date: str
    amount: str
    description: str
    date_format: str | None = None  # e.g. %Y-%m-%d, %Y%m%d, auto
    amount_decimal: str = ","  # Swedish often uses comma; "auto" detects per value
    delimiter: str | None = None


def sniff_csv(text: str) -> tuple[list[str], list[dict[str, str]], str]:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = list(reader.fieldnames or [])
    rows = []
    for i, row in enumerate(reader):
        rows.append({k: (v or "").strip() for k, v in row.items() if k is not None})
        if i >= 24:
            break
    return headers, rows, delimiter


def _cell_to_display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        # Keep fractional öre visible without scientific notation.
        return format(value, "f").rstrip("0").rstrip(".") if value != 0 else "0"
    return str(value).strip()


def read_tabular(path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            return [], [], ","
        headers = [
            str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(raw_rows[0])
        ]
        preview = []
        for row in raw_rows[1:26]:
            preview.append(
                {
                    headers[i]: "" if (i >= len(row)) else _cell_to_display(row[i])
                    for i in range(len(headers))
                }
            )
        return headers, preview, "xlsx"

    text = path.read_text(encoding="utf-8-sig", errors="replace")
    return sniff_csv(text)


def _clean_amount_text(raw: str) -> str:
    # Strip currency before collapsing spaces so "39,00 SEK" still matches.
    s = _CURRENCY_RE.sub("", raw.strip()).replace("\u00a0", " ").replace(" ", "")
    # Parentheses negative: (39,00)
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    return s


def _fraction_len(s: str, sep: str) -> int | None:
    if sep not in s:
        return None
    frac = s.rsplit(sep, 1)[-1]
    # Ignore trailing sign leftovers; amounts should be digits only after sep.
    if not frac.isdigit():
        return None
    return len(frac)


def detect_amount_decimal(raw: str, hint: str | None = ",") -> str:
    """Pick decimal separator from the number's shape.

    Clear shapes (1–2 fraction digits, or both separators present) win over a
    wrong UI hint — that is what turned Swedish ``-39,00`` into ``-3900``.
    """
    s = _clean_amount_text(raw)
    if not s or s in {"-", "+"}:
        return "," if hint in (None, "", "auto") else hint

    has_comma = "," in s
    has_dot = "." in s

    if has_comma and has_dot:
        return "," if s.rfind(",") > s.rfind(".") else "."

    if has_comma and not has_dot:
        fl = _fraction_len(s, ",")
        if fl is not None and fl <= 2:
            return ","
        if fl == 3:
            # Ambiguous thousands vs decimal — trust hint when provided.
            if hint in {".", ","}:
                return hint
            return "."  # 1,234 → thousands
        return ","

    if has_dot and not has_comma:
        fl = _fraction_len(s, ".")
        if fl is not None and fl <= 2:
            return "."
        if fl == 3:
            if hint in {".", ","}:
                return hint
            return ","  # 1.234 → thousands (European)
        return "."

    return "," if hint in (None, "", "auto") else hint


def guess_amount_decimal(samples: list[str], default: str = ",") -> str:
    """Majority-vote decimal separator from sample amount cells."""
    comma_votes = 0
    dot_votes = 0
    for raw in samples:
        if raw is None or str(raw).strip() == "":
            continue
        try:
            detected = detect_amount_decimal(str(raw), hint=default)
        except Exception:
            continue
        if detected == ",":
            comma_votes += 1
        else:
            dot_votes += 1
    if comma_votes == 0 and dot_votes == 0:
        return default
    return "," if comma_votes >= dot_votes else "."


def _parse_amount(raw: str, decimal: str = ",") -> float:
    s = _clean_amount_text(raw)
    if not s:
        raise ValueError("empty amount")

    # Always resolve separator from the value; wrong hints must not 100× amounts.
    sep = detect_amount_decimal(s, hint=decimal if decimal not in (None, "", "auto") else ",")
    if sep == ",":
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")
    return float(s)


def _coerce_amount(raw: Any, decimal: str = ",") -> float:
    if isinstance(raw, bool):
        raise ValueError("invalid amount")
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, Decimal):
        return float(raw)
    return _parse_amount(str(raw), decimal)


def _parse_date(raw: str, fmt: str | None) -> str:
    s = raw.strip()
    if not s:
        raise ValueError("empty date")
    formats = []
    if fmt and fmt != "auto":
        formats.append(fmt)
    formats.extend(
        [
            "%Y-%m-%d",
            "%Y%m%d",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%d.%m.%Y",
            "%Y/%m/%d",
            "%d/%m/%y",
            "%m/%d/%Y",
        ]
    )
    # Excel serial sometimes comes as float string
    try:
        as_float = float(s)
        if 30000 < as_float < 60000:
            from datetime import timedelta

            base = date(1899, 12, 30)
            return (base + timedelta(days=int(as_float))).isoformat()
    except ValueError:
        pass

    for f in formats:
        try:
            return datetime.strptime(s[:19], f).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"unparseable date: {raw!r}")


def _coerce_date(raw: Any, fmt: str | None) -> str:
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    return _parse_date(_cell_to_display(raw) if not isinstance(raw, str) else raw, fmt)


def parse_all_rows(path: Path, mapping: ColumnMapping) -> list[ParsedRow]:
    suffix = path.suffix.lower()
    rows_dicts: list[dict[str, Any]] = []

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers: list[str] = []
        row_count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [
                    str(c).strip() if c is not None else f"col{j}" for j, c in enumerate(row)
                ]
                continue
            d = {
                headers[j]: None if (j >= len(row) or row[j] is None) else row[j]
                for j in range(len(headers))
            }
            if any(v is not None and str(v).strip() != "" for v in d.values()):
                row_count += 1
                if row_count > MAX_IMPORT_ROWS:
                    raise ValueError(f"Too many rows (max {MAX_IMPORT_ROWS:,})")
                rows_dicts.append(d)
    else:
        text = path.read_text(encoding="utf-8-sig", errors="replace")
        delimiter = mapping.delimiter
        if not delimiter:
            _, _, delimiter = sniff_csv(text)
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        for row in reader:
            d = {k: (v or "").strip() for k, v in row.items() if k is not None}
            if any(d.values()):
                if len(rows_dicts) >= MAX_IMPORT_ROWS:
                    raise ValueError(f"Too many rows (max {MAX_IMPORT_ROWS:,})")
                rows_dicts.append(d)

    parsed: list[ParsedRow] = []
    for d in rows_dicts:
        try:
            tx_date = _coerce_date(d.get(mapping.date, ""), mapping.date_format)
            amount = _coerce_amount(d.get(mapping.amount, ""), mapping.amount_decimal)
            desc_raw = d.get(mapping.description, "")
            desc = _cell_to_display(desc_raw).strip() or "(no description)"
            parsed.append(ParsedRow(tx_date=tx_date, amount=amount, description=desc))
        except (ValueError, TypeError, KeyError):
            continue
    return parsed


def guess_mapping(headers: list[str]) -> dict[str, str | None]:
    lower = {h: h.lower() for h in headers}

    def find(*needles: str) -> str | None:
        # Prefer earlier needles across all headers (beskrivning before mottagare).
        for n in needles:
            for h, lh in lower.items():
                if n in lh:
                    return h
        return None

    return {
        "date": find(
            "bokföringsdag",
            "bokforingsdag",
            "transaktionsdag",
            "datum",
            "date",
            "valutadag",
        ),
        "amount": find("belopp", "amount", "summa"),
        # Prefer explicit description columns over counterparty fields.
        "description": find(
            "beskrivning",
            "text",
            "description",
            "meddelande",
            "narrativ",
            "mottagare",
            "avsändare",
            "avsandare",
        ),
    }
